"""Azure AI model batch deployment service.

Handles pre-check (plan), deployment execution with concurrent control,
and Redis-backed progress tracking (TTL 2h, no database).
"""

import asyncio
import io
import json
import logging
from uuid import uuid4

import redis.asyncio as aioredis
from azure.core.credentials import AccessToken
from azure.mgmt.cognitiveservices import CognitiveServicesManagementClient
from azure.mgmt.cognitiveservices.models import (
    Deployment,
    DeploymentModel,
    DeploymentProperties,
    Sku,
)

from app.config import settings
from app.services import azure_resource_service

logger = logging.getLogger(__name__)

TASK_TTL = 7200       # 2 hours
TOKEN_TTL = 3000      # ~50 minutes (ARM token valid ~1h)
CONCURRENCY = 3       # max parallel deployments per task
MAX_RETRIES = 3       # for Azure API throttling (429)
RETRY_BASE_DELAY = 5  # seconds


# ------------------------------------------------------------------ #
#  Redis singleton
# ------------------------------------------------------------------ #

_redis: aioredis.Redis | None = None


async def _get_redis() -> aioredis.Redis:
    global _redis
    if _redis is not None:
        try:
            await _redis.ping()
            return _redis
        except Exception:
            try:
                await _redis.aclose()
            except Exception:
                pass
            _redis = None

    url = settings.REDIS_URL.split("?")[0]
    _redis = aioredis.from_url(url, decode_responses=True, ssl_cert_reqs="none")
    await _redis.ping()
    return _redis


# ------------------------------------------------------------------ #
#  Credential wrapper
# ------------------------------------------------------------------ #

class _UserTokenCredential:
    def __init__(self, access_token: str, expires_on: int):
        self._token = AccessToken(access_token, expires_on)

    def get_token(self, *scopes, **kwargs) -> AccessToken:
        return self._token


# ------------------------------------------------------------------ #
#  Redis task helpers
# ------------------------------------------------------------------ #

async def _save_task(task_id: str, data: dict):
    r = await _get_redis()
    await r.setex(f"azure_deploy:{task_id}", TASK_TTL, json.dumps(data))


async def _load_task(task_id: str) -> dict | None:
    r = await _get_redis()
    raw = await r.get(f"azure_deploy:{task_id}")
    return json.loads(raw) if raw else None


async def _save_token(task_id: str, arm_token: str, expires_on: int):
    r = await _get_redis()
    await r.setex(
        f"azure_deploy_token:{task_id}",
        TOKEN_TTL,
        json.dumps({"arm_token": arm_token, "expires_on": expires_on}),
    )


async def _load_token(task_id: str) -> dict | None:
    r = await _get_redis()
    raw = await r.get(f"azure_deploy_token:{task_id}")
    return json.loads(raw) if raw else None


# ================================================================== #
#  Plan (pre-check)
# ================================================================== #

async def plan(arm_token: str, expires_on: int,
               subscription_id: str, items: list[dict]) -> dict:
    """Validate each deployment item against the target AI Foundry account.

    Checks: account existence, duplicate deployments, model availability,
    SKU compatibility. Returns action per item.
    """
    cred = _UserTokenCredential(arm_token, expires_on)
    cs_client = CognitiveServicesManagementClient(cred, subscription_id)

    # Cache existing deployments and available models per (rg, account, region)
    _deployment_cache: dict[tuple[str, str], dict[str, dict]] = {}
    _model_cache: dict[str, dict[str, dict]] = {}

    def _get_existing(rg: str, account: str) -> dict[str, dict]:
        key = (rg, account)
        if key not in _deployment_cache:
            try:
                deps = {}
                for d in cs_client.deployments.list(rg, account):
                    model_name = ""
                    model_version = ""
                    if d.properties and d.properties.model:
                        model_name = d.properties.model.name or ""
                        model_version = d.properties.model.version or ""
                    deps[d.name] = {
                        "model_name": model_name,
                        "model_version": model_version,
                    }
                _deployment_cache[key] = deps
            except Exception as e:
                logger.warning("Failed to list deployments for %s/%s: %s", rg, account, e)
                _deployment_cache[key] = {}
        return _deployment_cache[key]

    def _get_models(region: str) -> dict[str, dict]:
        if region not in _model_cache:
            try:
                models = {}
                for m in cs_client.models.list(location=region):
                    if m.model:
                        raw_skus = getattr(m, "skus", None) or getattr(m.model, "skus", None) or []
                        skus = [getattr(s, "name", "") for s in raw_skus if getattr(s, "name", None)]
                        model_format = getattr(m.model, "format", "") or ""
                        models[m.model.name] = {
                            "version": m.model.version or "",
                            "format": model_format,
                            "skus": skus,
                        }
                _model_cache[region] = models
            except Exception as e:
                logger.warning("Failed to list models for region %s: %s", region, e)
                _model_cache[region] = {}
        return _model_cache[region]

    plan_items = []
    counters = {"can_create": 0, "will_skip": 0, "has_conflict": 0,
                "unavailable": 0, "quota_risk": 0}

    for idx, item in enumerate(items):
        rg = item["resource_group"]
        account = item["account_name"]
        region = item["region"]
        model_name = item["model_name"]
        model_version = item["model_version"]
        dep_name = item["deployment_name"]
        sku_name = item["sku_name"]

        # 1. Check existing deployments on the target account
        existing = _get_existing(rg, account)
        if dep_name in existing:
            ex = existing[dep_name]
            if ex["model_name"] == model_name and ex["model_version"] == model_version:
                plan_items.append({
                    "index": idx,
                    "action": "skip",
                    "message": "同名部署已存在，模型和版本一致",
                })
                counters["will_skip"] += 1
                continue
            else:
                plan_items.append({
                    "index": idx,
                    "action": "conflict",
                    "message": f"同名部署已存在但版本不同 (现有: {ex['model_name']} {ex['model_version']})",
                })
                counters["has_conflict"] += 1
                continue

        # 2. Check model availability in region
        available = _get_models(region)
        if model_name not in available:
            plan_items.append({
                "index": idx,
                "action": "unavailable",
                "message": f"模型 {model_name} 在区域 {region} 不可用",
            })
            counters["unavailable"] += 1
            continue

        # 3. Check SKU compatibility
        model_info = available[model_name]
        if model_info["skus"] and sku_name not in model_info["skus"]:
            plan_items.append({
                "index": idx,
                "action": "unavailable",
                "message": f"SKU {sku_name} 不支持该模型，可用: {', '.join(model_info['skus'])}",
            })
            counters["unavailable"] += 1
            continue

        # 4. Passed all checks
        plan_items.append({"index": idx, "action": "create", "message": None})
        counters["can_create"] += 1

    return {
        "total": len(items),
        **counters,
        "items": plan_items,
    }


# ================================================================== #
#  Start deployment
# ================================================================== #

async def start_deployment(arm_token: str, expires_on: int,
                           subscription_id: str,
                           items: list[dict]) -> str:
    """Create a task record in Redis and return the task_id.

    The caller (API route) should then kick off ``execute()`` in a
    BackgroundTask.
    """
    task_id = uuid4().hex[:8]
    task_data = {
        "status": "running",
        "total": len(items),
        "succeeded": 0,
        "failed": 0,
        "subscription_id": subscription_id,
        "items": [
            {
                "index": i,
                "model_name": item["model_name"],
                "model_version": item.get("model_version", ""),
                "model_format": item.get("model_format", "OpenAI"),
                "region": item["region"],
                "account_name": item["account_name"],
                "resource_group": item["resource_group"],
                "deployment_name": item["deployment_name"],
                "sku_name": item.get("sku_name", "GlobalStandard"),
                "sku_capacity": item.get("sku_capacity", 10),
                "status": "pending",
                "error": None,
            }
            for i, item in enumerate(items)
        ],
    }

    await _save_task(task_id, task_data)
    await _save_token(task_id, arm_token, expires_on)
    return task_id


# ================================================================== #
#  Execute deployment (called from BackgroundTask)
# ================================================================== #

async def execute(task_id: str):
    """Run all pending items, with concurrency control and retry on 429."""
    task = await _load_task(task_id)
    token_data = await _load_token(task_id)
    if not task or not token_data:
        logger.error("Task %s: missing task or token data", task_id)
        return

    cred = _UserTokenCredential(token_data["arm_token"], token_data["expires_on"])
    sem = asyncio.Semaphore(CONCURRENCY)

    async def deploy_one(idx: int, item: dict):
        async with sem:
            item["status"] = "deploying"
            await _save_task(task_id, task)

            try:
                await _create_deployment_with_retry(
                    cred,
                    task["subscription_id"],
                    item,
                )
                item["status"] = "succeeded"
                task["succeeded"] += 1
            except Exception as e:
                item["status"] = "failed"
                item["error"] = str(e)
                task["failed"] += 1
                logger.error("Task %s item %d failed: %s", task_id, idx, e)

            await _save_task(task_id, task)

    pending = [
        (i, item) for i, item in enumerate(task["items"])
        if item["status"] == "pending"
    ]
    if pending:
        await asyncio.gather(*(deploy_one(i, item) for i, item in pending))

    task["status"] = "completed"
    await _save_task(task_id, task)


async def _create_deployment_with_retry(
    credential: _UserTokenCredential,
    subscription_id: str,
    item: dict,
):
    """Create a single deployment, retrying on 429 (TooManyRequests)."""
    cs_client = CognitiveServicesManagementClient(credential, subscription_id)

    model_format = item.get("model_format", "OpenAI") or "OpenAI"
    deployment = Deployment(
        sku=Sku(name=item["sku_name"], capacity=item["sku_capacity"]),
        properties=DeploymentProperties(
            model=DeploymentModel(
                format=model_format,
                name=item["model_name"],
                version=item["model_version"],
            )
        ),
    )

    last_err: Exception | None = None
    for attempt in range(MAX_RETRIES):
        try:
            poller = cs_client.deployments.begin_create_or_update(
                resource_group_name=item["resource_group"],
                account_name=item["account_name"],
                deployment_name=item["deployment_name"],
                deployment=deployment,
            )
            poller.result()
            return
        except Exception as e:
            last_err = e
            err_str = str(e).lower()
            if "429" in err_str or "toomanyrequests" in err_str or "throttl" in err_str:
                delay = RETRY_BASE_DELAY * (2 ** attempt)
                logger.warning(
                    "Throttled on %s (attempt %d/%d), retrying in %ds",
                    item["deployment_name"], attempt + 1, MAX_RETRIES, delay,
                )
                await asyncio.sleep(delay)
                continue
            raise

    raise last_err  # type: ignore[misc]


# ================================================================== #
#  Progress
# ================================================================== #

async def get_progress(task_id: str) -> dict | None:
    task = await _load_task(task_id)
    if not task:
        return None

    items = task["items"]
    status_counts = {"pending": 0, "deploying": 0, "succeeded": 0, "failed": 0}
    for item in items:
        s = item.get("status", "pending")
        if s in status_counts:
            status_counts[s] += 1

    return {
        "task_id": task_id,
        "status": task["status"],
        "total": task["total"],
        "succeeded": status_counts["succeeded"],
        "failed": status_counts["failed"],
        "deploying": status_counts["deploying"],
        "pending": status_counts["pending"],
        "items": [
            {
                "index": item["index"],
                "model_name": item["model_name"],
                "region": item["region"],
                "account_name": item["account_name"],
                "deployment_name": item["deployment_name"],
                "status": item["status"],
                "error": item.get("error"),
            }
            for item in items
        ],
    }


# ================================================================== #
#  Retry failed items
# ================================================================== #

async def retry_failed(task_id: str, arm_token: str, expires_on: int) -> int:
    """Reset failed items to pending, update the ARM token, return retry count."""
    task = await _load_task(task_id)
    if not task:
        raise ValueError("任务不存在或已过期")

    retrying = 0
    for item in task["items"]:
        if item["status"] == "failed":
            item["status"] = "pending"
            item["error"] = None
            retrying += 1

    if retrying == 0:
        raise ValueError("没有失败项可重试")

    task["status"] = "running"
    task["failed"] = 0
    await _save_task(task_id, task)
    await _save_token(task_id, arm_token, expires_on)
    return retrying


# ================================================================== #
#  Excel export (deployment name, endpoint, key, RPM, TPM)
# ================================================================== #


async def build_export_xlsx_bytes(
    arm_token: str,
    expires_on: int,
    task_id: str,
) -> tuple[bytes, str]:
    """Build an .xlsx from Redis task data + live Azure endpoint/keys/quota."""
    task = await _load_task(task_id)
    if not task:
        raise ValueError("任务不存在或已过期")

    sub_id = task.get("subscription_id")
    items = task.get("items") or []
    if not sub_id or not items:
        raise ValueError("任务数据不完整")

    rows = azure_resource_service.fetch_deploy_export_rows(
        arm_token, expires_on, sub_id, items,
    )

    from openpyxl import Workbook
    from openpyxl.styles import Alignment, Font, PatternFill

    wb = Workbook()
    ws = wb.active
    ws.title = "部署清单"

    headers = ["部署名", "终结点", "密钥", "RPM", "TPM"]
    header_font = Font(bold=True, color="FFFFFF", size=11)
    header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
    header_align = Alignment(horizontal="center")

    for ci, h in enumerate(headers, 1):
        cell = ws.cell(row=1, column=ci, value=h)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_align

    for ri, row in enumerate(rows, 2):
        ws.cell(row=ri, column=1, value=row["deployment_name"])
        ws.cell(row=ri, column=2, value=row["endpoint"])
        ws.cell(row=ri, column=3, value=row["api_key"])
        ws.cell(row=ri, column=4, value=row["rpm"] if row["rpm"] is not None else "")
        ws.cell(row=ri, column=5, value=row["tpm"] if row["tpm"] is not None else "")

    ws.column_dimensions["A"].width = 28
    ws.column_dimensions["B"].width = 72
    ws.column_dimensions["C"].width = 48
    ws.column_dimensions["D"].width = 12
    ws.column_dimensions["E"].width = 14

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    fname = f"azure-deploy-{task_id}.xlsx"
    return buf.getvalue(), fname
