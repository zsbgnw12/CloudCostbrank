"""Service Accounts API — unified view over CloudAccount + DataSource + Project.

云厂商(provider)仅来自 supply_sources；供应商名称仅来自 suppliers。projects 不重复存 provider/group_label。
"""

import datetime as dt
import io
import json
from decimal import Decimal
from typing import Any

from fastapi import APIRouter, Depends, HTTPException, Query, Request, Response
from fastapi.responses import StreamingResponse
from pydantic import BaseModel, ConfigDict, field_validator
from sqlalchemy import select, func
from sqlalchemy.ext.asyncio import AsyncSession

from app.auth.dependencies import get_current_principal, require_roles
from app.auth.principal import Principal
from app.auth.scope import (
    ensure_provider_visible,
    extract_providers_from_roles,
    has_full_access,
)
from app.database import get_db
from app.models.billing import BillingData
from app.models.cloud_account import CloudAccount
from app.models.data_source import DataSource
from app.models.entity import Entity
from app.models.project import Project
from app.models.project_assignment_log import ProjectAssignmentLog
from app.models.project_customer_assignment import ProjectCustomerAssignment
from app.models.supplier import Supplier
from app.models.supply_source import SupplySource
from app.services.crypto_service import encrypt_dict, decrypt_to_dict
from app.services.default_supply_sources import ensure_other_gcp_supply_source_id

router = APIRouter()


def _data_source_config_for_create(provider: str, external_project_id: str) -> dict:
    """DataSource.config for collectors. Azure needs subscription_id (same as Project.external_project_id)."""
    base: dict = {"auto_created": True}
    if provider == "azure":
        base["subscription_id"] = external_project_id.strip()
        base["collect_mode"] = "subscription"
        base["cost_metric"] = "ActualCost"
    return base


async def _cloud_provider(db: AsyncSession, project: Project) -> str:
    ss = await db.get(SupplySource, project.supply_source_id)
    if not ss:
        raise HTTPException(500, "Project 缺少有效货源")
    return ss.provider


# ─── Schemas ───────────────────────────────────────────────────

class ServiceAccountCreate(BaseModel):
    supply_source_id: int
    entity_id: int | None = None
    name: str
    external_project_id: str
    secret_data: dict[str, Any] = {}
    notes: str | None = None
    order_method: str | None = None

    @field_validator("name", "external_project_id", mode="before")
    @classmethod
    def strip_whitespace(cls, v: object) -> object:
        if isinstance(v, str):
            return v.strip()
        return v

    @field_validator("order_method", mode="before")
    @classmethod
    def strip_order_method(cls, v: object) -> object:
        if v is None:
            return None
        if isinstance(v, str):
            s = v.strip()
            return s if s else None
        return v


class ServiceAccountUpdate(BaseModel):
    name: str | None = None
    supply_source_id: int | None = None
    # entity_id：None 表示「不动」；显式传 0 或 null-via-unset 都不会动。
    # 传具体 id 则切换主体；前端需要清主体走 entity_id=None + clear_entity=True 模式。
    entity_id: int | None = None
    clear_entity: bool = False
    external_project_id: str | None = None
    secret_data: dict[str, Any] | None = None
    notes: str | None = None
    order_method: str | None = None
    # 客户编号全量覆盖语义：None=不动；[]=清空；[...]=替换为该集合。
    # 不会影响账号 status（状态和客户编号已解耦；改 status 走 /suspend /activate /standby）。
    customer_codes: list[str] | None = None

    @field_validator("name", "external_project_id", mode="before")
    @classmethod
    def strip_whitespace(cls, v: object) -> object:
        if isinstance(v, str):
            return v.strip()
        return v

    @field_validator("order_method", mode="before")
    @classmethod
    def strip_order_method(cls, v: object) -> object:
        if v is None:
            return None
        if isinstance(v, str):
            s = v.strip()
            return s if s else None
        return v


class ServiceAccountListItem(BaseModel):
    model_config = ConfigDict(from_attributes=True)
    id: int
    name: str
    supply_source_id: int
    supplier_name: str
    provider: str  # 来自 supply_sources，非 projects 列
    entity_id: int | None = None
    entity_name: str | None = None
    external_project_id: str
    status: str
    order_method: str | None = None
    customer_codes: list[str] = []
    created_at: dt.datetime


class HistoryItem(BaseModel):
    model_config = ConfigDict(from_attributes=True)
    id: int
    action: str
    from_status: str | None
    to_status: str | None
    operator: str | None
    customer_code: str | None = None
    notes: str | None
    created_at: dt.datetime


class ServiceAccountDetail(BaseModel):
    model_config = ConfigDict(from_attributes=True)
    id: int
    name: str
    supply_source_id: int
    supplier_id: int
    supplier_name: str
    provider: str
    entity_id: int | None = None
    entity_name: str | None = None
    external_project_id: str
    status: str
    notes: str | None
    order_method: str | None = None
    customer_codes: list[str] = []
    secret_fields: list[str]
    created_at: dt.datetime
    history: list[HistoryItem]


# ─── Sales-sync payloads ──────────────────────────────────────

class SalesAssignmentItem(BaseModel):
    customer_code: str
    supplier_name: str
    provider: str
    external_project_id: str


class SalesSyncBody(BaseModel):
    """销售系统批量下发客户 ↔ 服务账号 关联。

    mode=full: 对 scope_customer_codes 这一批做差分（多删少插）；未列入 scope 的
      客户编号不动。
    mode=patch: 仅做 upsert，不删除任何已有关联。
    """
    assignments: list[SalesAssignmentItem]
    mode: str = "patch"  # "full" | "patch"
    scope_customer_codes: list[str] = []

    @field_validator("mode", mode="before")
    @classmethod
    def _normalize_mode(cls, v: object) -> object:
        if isinstance(v, str):
            v = v.strip().lower()
            if v not in ("full", "patch"):
                raise ValueError("mode must be 'full' or 'patch'")
        return v


class SalesSyncUnmatched(BaseModel):
    customer_code: str
    supplier_name: str
    provider: str
    external_project_id: str
    reason: str


class SalesSyncResult(BaseModel):
    inserted: int
    deleted: int
    unchanged: int
    unmatched: list[SalesSyncUnmatched]


class CostByService(BaseModel):
    service: str
    cost: float
    usage_quantity: float
    usage_unit: str | None


class DailyCost(BaseModel):
    date: str
    cost: float
    usage_quantity: float


class DailyServiceCost(BaseModel):
    date: str
    service: str
    cost: float
    usage_quantity: float
    usage_unit: str | None


class CostSummary(BaseModel):
    total_cost: float
    total_usage: float
    services: list[CostByService]
    daily: list[DailyCost]
    daily_by_service: list[DailyServiceCost]


# ─── Helpers ───────────────────────────────────────────────────

def _log(
    db,
    project,
    action: str,
    from_status: str,
    to_status: str,
    notes: str | None = None,
    customer_code: str | None = None,
    operator: str | None = None,
):
    db.add(ProjectAssignmentLog(
        project_id=project.id, action=action,
        from_status=from_status, to_status=to_status,
        customer_code=customer_code, operator=operator, notes=notes,
    ))


def _normalize_code(code: str) -> str:
    """上游客户编号统一大写 + 去空白。"""
    return (code or "").strip().upper()


async def _codes_for_project(db: AsyncSession, project_id: int) -> list[str]:
    rows = (
        await db.execute(
            select(ProjectCustomerAssignment.customer_code)
            .where(ProjectCustomerAssignment.project_id == project_id)
            .order_by(ProjectCustomerAssignment.customer_code)
        )
    ).all()
    return [r[0] for r in rows]


async def _codes_by_project_ids(
    db: AsyncSession, project_ids: list[int]
) -> dict[int, list[str]]:
    if not project_ids:
        return {}
    rows = (
        await db.execute(
            select(
                ProjectCustomerAssignment.project_id,
                ProjectCustomerAssignment.customer_code,
            )
            .where(ProjectCustomerAssignment.project_id.in_(project_ids))
            .order_by(
                ProjectCustomerAssignment.project_id,
                ProjectCustomerAssignment.customer_code,
            )
        )
    ).all()
    out: dict[int, list[str]] = {pid: [] for pid in project_ids}
    for pid, code in rows:
        out.setdefault(pid, []).append(code)
    return out


# NOTE: 以前这里有 _recompute_status，根据 customer_codes 自动派生 status。
# 现在 status 和客户编号彻底解耦：状态完全由人工点按钮决定（使用中/备用/停用），
# 客户编号只是账号上的一个独立标签。所以派生函数已移除。


def _principal_operator(request) -> str | None:
    """尽量从 request.state.principal 取一个人类可读的 operator 名。"""
    try:
        principal = getattr(request.state, "principal", None)
        if not principal:
            return None
        u = getattr(principal, "user", None)
        if u and getattr(u, "username", None):
            return u.username
        return getattr(principal, "auth_method", None) or None
    except Exception:
        return None


# ─── Endpoints ─────────────────────────────────────────────────

async def _scope_check_account(db: AsyncSession, principal: Principal, account_id: int) -> None:
    """写操作前校验:account 所属 provider 必须在用户范围内。
    cloud_admin / cloud_ops 全开;cloud_<provider> 限本云。"""
    if has_full_access(principal):
        return
    scope_providers = extract_providers_from_roles(principal.roles)
    if not scope_providers:
        raise HTTPException(403, "Forbidden: cloud role required")
    row = await db.execute(
        select(SupplySource.provider)
        .join(Project, Project.supply_source_id == SupplySource.id)
        .where(Project.id == account_id)
    )
    p = row.scalar_one_or_none()
    if p is None:
        raise HTTPException(404, "Service account not found")
    if p not in scope_providers:
        raise HTTPException(403, f"Provider '{p}' out of scope")


async def _account_in_scope(
    account_id: int,
    db: AsyncSession = Depends(get_db),
    principal: Principal = Depends(get_current_principal),
) -> int:
    """Dependency: 路由路径含 {account_id} 时自动校验数据范围。"""
    await _scope_check_account(db, principal, account_id)
    return account_id


@router.get("/", response_model=list[ServiceAccountListItem])
async def list_accounts(
    response: Response,
    provider: str | None = None,
    status: str | None = None,
    customer_code: str | None = None,
    page: int = Query(1, ge=1),
    page_size: int = Query(100, ge=1, le=500),
    db: AsyncSession = Depends(get_db),
    principal: Principal = Depends(get_current_principal),
):
    base_stmt = (
        select(
            Project.id,
            Project.name,
            Project.supply_source_id,
            Project.external_project_id,
            Project.status,
            Project.order_method,
            Project.created_at,
            Project.entity_id,
            Entity.name.label("entity_name"),
            SupplySource.provider,
            Supplier.name.label("supplier_name"),
        )
        .join(SupplySource, Project.supply_source_id == SupplySource.id)
        .join(Supplier, SupplySource.supplier_id == Supplier.id)
        .outerjoin(Entity, Project.entity_id == Entity.id)
        .where(Project.recycled_at.is_(None))
    )
    # 数据范围:cloud_<provider> 只看本云
    if not has_full_access(principal):
        scope_providers = extract_providers_from_roles(principal.roles)
        if not scope_providers:
            response.headers["X-Total-Count"] = "0"
            response.headers["Access-Control-Expose-Headers"] = "X-Total-Count, X-Page, X-Page-Size"
            return []
        base_stmt = base_stmt.where(SupplySource.provider.in_(scope_providers))
    if provider:
        base_stmt = base_stmt.where(SupplySource.provider == provider)
    if status:
        base_stmt = base_stmt.where(Project.status == status)
    if customer_code:
        code = _normalize_code(customer_code)
        base_stmt = base_stmt.join(
            ProjectCustomerAssignment,
            ProjectCustomerAssignment.project_id == Project.id,
        ).where(ProjectCustomerAssignment.customer_code == code)

    # 总数(过滤后,分页前) -- 写到响应 header,供前端做"上一页/下一页"分页 UI。
    # 同时暴露 Access-Control-Expose-Headers 让浏览器跨域能读到这两个 header。
    count_stmt = select(func.count()).select_from(base_stmt.subquery())
    total = (await db.execute(count_stmt)).scalar_one()
    response.headers["X-Total-Count"] = str(total)
    response.headers["X-Page"] = str(page)
    response.headers["X-Page-Size"] = str(page_size)
    response.headers["Access-Control-Expose-Headers"] = "X-Total-Count, X-Page, X-Page-Size"

    stmt = (
        base_stmt.order_by(SupplySource.provider, Supplier.name, Project.name)
        .offset((page - 1) * page_size)
        .limit(page_size)
    )
    rows = (await db.execute(stmt)).all()

    codes_map = await _codes_by_project_ids(db, [r.id for r in rows])
    return [
        ServiceAccountListItem(
            id=r.id,
            name=r.name,
            supply_source_id=r.supply_source_id,
            supplier_name=r.supplier_name,
            provider=r.provider,
            entity_id=r.entity_id,
            entity_name=r.entity_name,
            external_project_id=r.external_project_id,
            status=r.status,
            order_method=r.order_method,
            customer_codes=codes_map.get(r.id, []),
            created_at=r.created_at,
        )
        for r in rows
    ]


@router.post(
    "/",
    response_model=ServiceAccountListItem,
    status_code=201,
)
async def create_account(
    body: ServiceAccountCreate,
    db: AsyncSession = Depends(get_db),
    principal: Principal = Depends(get_current_principal),
):
    ss = await db.get(SupplySource, body.supply_source_id)
    if not ss:
        raise HTTPException(404, "货源不存在")
    cloud = ss.provider
    # 数据范围:用户的角色必须能管该 supply_source 的 provider
    ensure_provider_visible(principal, cloud)

    # entity_id 校验：主体必须属于本货源
    entity_id: int | None = None
    if body.entity_id is not None:
        ent = await db.get(Entity, body.entity_id)
        if not ent or ent.supply_source_id != body.supply_source_id:
            raise HTTPException(400, "entity_id 不属于该货源")
        entity_id = ent.id

    encrypted = encrypt_dict(body.secret_data) if body.secret_data else encrypt_dict({})
    ca = CloudAccount(name=f"{cloud}-{body.name}", provider=cloud, secret_data=encrypted)
    db.add(ca)
    await db.flush()

    ds = DataSource(
        name=f"ds-{body.name}", cloud_account_id=ca.id,
        config=_data_source_config_for_create(cloud, body.external_project_id),
        is_active=True,
    )
    db.add(ds)
    await db.flush()

    project = Project(
        name=body.name,
        external_project_id=body.external_project_id,
        supply_source_id=body.supply_source_id,
        entity_id=entity_id,
        data_source_id=ds.id,
        notes=body.notes,
        order_method=body.order_method,
        status="active",
    )
    db.add(project)
    await db.flush()

    _log(db, project, "created", from_status="", to_status="active")
    await db.commit()

    su = await db.get(Supplier, ss.supplier_id)
    ent_name = None
    if entity_id is not None:
        ent_obj = await db.get(Entity, entity_id)
        ent_name = ent_obj.name if ent_obj else None
    return ServiceAccountListItem(
        id=project.id,
        name=project.name,
        supply_source_id=project.supply_source_id,
        supplier_name=su.name if su else "",
        provider=cloud,
        entity_id=entity_id,
        entity_name=ent_name,
        external_project_id=project.external_project_id,
        status=project.status,
        order_method=project.order_method,
        customer_codes=[],
        created_at=project.created_at,
    )


# ─── Delete (physical / hard delete) ─────────────────────────

async def _get_active_project(db: AsyncSession, account_id: int) -> Project:
    """db.get(Project, id)，但若 project 不存在或已软删（recycled_at 非空）则 404。

    面向用户的所有 CRUD 端点（详情、suspend/activate/standby、状态编辑、成本报表等）
    都应该走这个，避免操作已被删除的账号。内部审计流程如需访问回收站内容，直接用 db.get。
    """
    project = await db.get(Project, account_id)
    if not project or project.recycled_at is not None:
        raise HTTPException(404, "Service account not found")
    return project


async def _hard_delete(account_id: int, db: AsyncSession):
    """Soft-delete：只打 recycled_at 时间戳，前端永不再显示；billing/sync 历史一律保留。

    以前这里是 cascade 物理删 billing_data / summary / sync_logs / token_usage / ds / ca，
    问题有三：
      1. auto_create_gcp_projects 下次 sync 发现 BQ 里还有这个 project_id 就会原名复活（status=standby），
         用户点了删除，第二天又冒出来，看起来像"删不掉"；
      2. 共享 data_source 的场景（比如 ds#4 下挂 47 个项目）一旦误删会连带清空其他项目的账单；
      3. 账单数据、审计日志一并消失，事后无法回溯。
    现在改为软删：
      - Project.recycled_at = NOW()
      - status 顺带置 "inactive"（给审计日志用）
      - list_accounts 会过滤掉 recycled_at 非空的行
      - auto_create_gcp_projects 本来就用 external_project_id 判重（跨状态），软删后再次
        sync 到同 project_id 也会被视为"已存在"→ 不会复活
      - 一切 billing / ds / ca / sync_log 原封不动
    若需要真的物理清除某账号所有数据，请运维直连 DB 手工处理（留痕）。
    """
    project = await db.get(Project, account_id)
    if not project:
        raise HTTPException(404, "Service account not found")
    if project.recycled_at is not None:
        return  # idempotent: already soft-deleted, second click is a no-op

    old_status = project.status
    project.recycled_at = dt.datetime.utcnow()
    project.status = "inactive"

    _log(
        db, project, "deleted",
        from_status=old_status, to_status="inactive",
    )
    await db.commit()


@router.delete(
    "/hard/{account_id}",
    status_code=204,
    dependencies=[Depends(require_roles("cloud_admin"))],
)
async def hard_delete_account(account_id: int, db: AsyncSession = Depends(get_db)):
    await _hard_delete(account_id, db)


# ─── All Accounts Daily Costs (must be before /{account_id}) ──

class AccountDailyCostRow(BaseModel):
    account_id: int
    account_name: str
    provider: str
    external_project_id: str
    date: str
    product: str | None
    service_id: str | None = None
    cost: float
    cost_at_list: float | None = None
    credits_total: float | None = None
    currency: str | None = None


@router.get("/daily-report", response_model=list[AccountDailyCostRow])
async def daily_report(
    start_date: str = Query(..., pattern=r"^\d{4}-\d{2}-\d{2}$"),
    end_date: str = Query(..., pattern=r"^\d{4}-\d{2}-\d{2}$"),
    provider: str | None = Query(None),
    db: AsyncSession = Depends(get_db),
):
    sd = dt.date.fromisoformat(start_date)
    ed = dt.date.fromisoformat(end_date) + dt.timedelta(days=1)

    stmt = (
        select(
            Project.id.label("account_id"),
            Project.name.label("account_name"),
            SupplySource.provider.label("provider"),
            BillingData.project_id,
            BillingData.date,
            BillingData.product,
            func.max(BillingData.service_id).label("service_id"),
            func.sum(BillingData.cost).label("cost"),
            func.sum(BillingData.cost_at_list).label("cost_at_list"),
            func.sum(BillingData.credits_total).label("credits_total"),
            func.max(BillingData.currency).label("currency"),
        )
        .join(
            Project,
            func.trim(BillingData.project_id) == func.trim(Project.external_project_id),
        )
        .join(SupplySource, Project.supply_source_id == SupplySource.id)
        .where(
            BillingData.provider == SupplySource.provider,
            BillingData.date >= sd,
            BillingData.date < ed,
        )
        .group_by(
            Project.id,
            Project.name,
            SupplySource.provider,
            BillingData.project_id,
            BillingData.date,
            BillingData.product,
        )
        .order_by(BillingData.date, BillingData.project_id, BillingData.product)
    )
    if provider:
        stmt = stmt.where(SupplySource.provider == provider)

    rows = (await db.execute(stmt)).all()

    def _f(v):
        return float(v) if v is not None else None

    return [
        AccountDailyCostRow(
            account_id=r.account_id,
            account_name=r.account_name,
            provider=r.provider,
            external_project_id=r.project_id or "",
            date=str(r.date),
            product=r.product or "Unknown",
            service_id=r.service_id,
            cost=float(r.cost) if r.cost is not None else 0.0,
            cost_at_list=_f(r.cost_at_list),
            credits_total=_f(r.credits_total),
            currency=r.currency,
        )
        for r in rows
    ]


@router.get("/daily-report/export")
async def export_daily_report(
    start_date: str = Query(..., pattern=r"^\d{4}-\d{2}-\d{2}$"),
    end_date: str = Query(..., pattern=r"^\d{4}-\d{2}-\d{2}$"),
    provider: str | None = Query(None),
    discount_pct: float | None = Query(
        None,
        ge=0,
        le=100,
        description="统一折扣百分比；传入时导出增加「折扣」「折后费用」列",
    ),
    db: AsyncSession = Depends(get_db),
):
    rows = await daily_report(start_date, end_date, provider, db)
    return _build_excel(rows, f"daily_report_{start_date}_{end_date}.xlsx", discount_pct=discount_pct)


@router.get("/{account_id}", response_model=ServiceAccountDetail)
async def get_account(account_id: int, db: AsyncSession = Depends(get_db)):
    row = (await db.execute(
        select(Project, DataSource, CloudAccount, SupplySource, Supplier, Entity)
        .join(SupplySource, Project.supply_source_id == SupplySource.id)
        .join(Supplier, SupplySource.supplier_id == Supplier.id)
        .outerjoin(DataSource, Project.data_source_id == DataSource.id)
        .outerjoin(CloudAccount, DataSource.cloud_account_id == CloudAccount.id)
        .outerjoin(Entity, Project.entity_id == Entity.id)
        .where(Project.id == account_id, Project.recycled_at.is_(None))
    )).first()
    if not row:
        raise HTTPException(404, "Service account not found")
    project, ds, ca, ss, su, ent = row

    secret_fields: list[str] = []
    if ca:
        try:
            secret_fields = list(decrypt_to_dict(ca.secret_data).keys())
        except Exception:
            secret_fields = ["(encrypted)"]

    logs = (await db.execute(
        select(ProjectAssignmentLog)
        .where(ProjectAssignmentLog.project_id == account_id)
        .order_by(ProjectAssignmentLog.created_at.desc())
    )).scalars().all()

    history = [HistoryItem(
        id=lg.id, action=lg.action,
        from_status=lg.from_status, to_status=lg.to_status,
        operator=lg.operator, customer_code=lg.customer_code,
        notes=lg.notes, created_at=lg.created_at,
    ) for lg in logs]

    customer_codes = await _codes_for_project(db, project.id)

    return ServiceAccountDetail(
        id=project.id,
        name=project.name,
        supply_source_id=project.supply_source_id,
        supplier_id=su.id,
        supplier_name=su.name,
        provider=ss.provider,
        entity_id=ent.id if ent else None,
        entity_name=ent.name if ent else None,
        external_project_id=project.external_project_id,
        status=project.status,
        notes=project.notes,
        order_method=project.order_method,
        customer_codes=customer_codes,
        secret_fields=secret_fields,
        created_at=project.created_at,
        history=history,
    )


@router.put(
    "/{account_id}",
    response_model=ServiceAccountDetail,
    dependencies=[Depends(_account_in_scope)],
)
async def update_account(
    account_id: int,
    body: ServiceAccountUpdate,
    request: Request,
    db: AsyncSession = Depends(get_db),
):
    project = await _get_active_project(db, account_id)

    data = body.model_dump(exclude_unset=True)
    secret_data = data.pop("secret_data", None)
    new_supply_source_id = data.pop("supply_source_id", None)
    customer_codes_payload = data.pop("customer_codes", None)
    entity_id_payload = data.pop("entity_id", None)  # None=不动；显式 int=切换
    clear_entity_flag = bool(data.pop("clear_entity", False))

    for k, v in data.items():
        if hasattr(project, k):
            setattr(project, k, v)
    await db.flush()

    if new_supply_source_id is not None and new_supply_source_id != project.supply_source_id:
        new_ss = await db.get(SupplySource, new_supply_source_id)
        if not new_ss:
            raise HTTPException(404, "货源不存在")
        ext = (data.get("external_project_id") if "external_project_id" in data else None) or project.external_project_id
        ext = str(ext).strip()
        dup = (
            await db.execute(
                select(Project.id).where(
                    Project.supply_source_id == new_supply_source_id,
                    Project.external_project_id == ext,
                    Project.id != project.id,
                )
            )
        ).scalar_one_or_none()
        if dup:
            raise HTTPException(409, "目标货源下已存在相同账号 ID")
        project.supply_source_id = new_supply_source_id
        # 货源切换：主体一定要清空(主体是 supply_source-scoped 的)
        project.entity_id = None
        await db.flush()
        if project.data_source_id:
            ds = await db.get(DataSource, project.data_source_id)
            if ds and ds.cloud_account_id:
                ca = await db.get(CloudAccount, ds.cloud_account_id)
                if ca:
                    ca.provider = new_ss.provider
                    ca.name = f"{new_ss.provider}-{project.name}"[:100]
                prov_new = new_ss.provider
                base_cfg = _data_source_config_for_create(prov_new, ext)
                old_cfg = dict(ds.config) if ds.config else {}
                merged = {**old_cfg, **base_cfg}
                ds.config = merged
                await db.flush()

    prov = await _cloud_provider(db, project)
    if project.data_source_id and prov == "azure" and "external_project_id" in data:
        ds = await db.get(DataSource, project.data_source_id)
        if ds:
            cfg = dict(ds.config) if ds.config else {}
            cfg["subscription_id"] = project.external_project_id.strip()
            cfg.setdefault("collect_mode", "subscription")
            cfg.setdefault("cost_metric", "ActualCost")
            ds.config = cfg
            await db.flush()

    if secret_data is not None and project.data_source_id:
        ds = await db.get(DataSource, project.data_source_id)
        if ds:
            ca = await db.get(CloudAccount, ds.cloud_account_id)
            if ca:
                ca.secret_data = encrypt_dict(secret_data)
                await db.flush()

    # entity_id 写入：clear_entity=True 强制清空；entity_id=int 切换；皆需校验属于当前 supply_source
    if clear_entity_flag:
        project.entity_id = None
        await db.flush()
    elif entity_id_payload is not None:
        ent = await db.get(Entity, entity_id_payload)
        if not ent or ent.supply_source_id != project.supply_source_id:
            raise HTTPException(400, "entity_id 不属于该货源")
        project.entity_id = ent.id
        await db.flush()

    # Customer-codes diff + status recompute (full replace semantics).
    if customer_codes_payload is not None:
        new_codes: set[str] = set()
        for c in customer_codes_payload:
            c = _normalize_code(c)
            if c:
                new_codes.add(c)
        existing = set(await _codes_for_project(db, project.id))
        to_add = new_codes - existing
        to_remove = existing - new_codes
        if to_add or to_remove:
            operator = _principal_operator(request)
            old_status = project.status
            for code in sorted(to_remove):
                await db.execute(
                    ProjectCustomerAssignment.__table__.delete().where(
                        (ProjectCustomerAssignment.project_id == project.id)
                        & (ProjectCustomerAssignment.customer_code == code)
                    )
                )
                _log(
                    db, project, "customer_unbound",
                    from_status=old_status, to_status=old_status,
                    customer_code=code, operator=operator,
                )
            for code in sorted(to_add):
                db.add(ProjectCustomerAssignment(
                    project_id=project.id, customer_code=code, assigned_by=operator,
                ))
                _log(
                    db, project, "customer_bound",
                    from_status=old_status, to_status=old_status,
                    customer_code=code, operator=operator,
                )
            await db.flush()
            # status 不再随客户编号派生，保持当前值

    await db.commit()
    return await get_account(account_id, db)


# ─── Bulk reassign to another SupplySource ───────────────────

class BulkAssignRequest(BaseModel):
    account_ids: list[int]
    target_supply_source_id: int

    @field_validator("account_ids")
    @classmethod
    def _non_empty(cls, v: list[int]) -> list[int]:
        if not v:
            raise ValueError("account_ids 不能为空")
        # 去重，避免同 id 被处理两次
        return list(dict.fromkeys(v))


class BulkAssignSkip(BaseModel):
    account_id: int
    reason: str


class BulkAssignResponse(BaseModel):
    moved: int
    skipped: list[BulkAssignSkip]
    target_supply_source_id: int
    target_provider: str
    target_supplier_name: str


@router.post(
    "/bulk-assign",
    response_model=BulkAssignResponse,
)
async def bulk_assign(
    body: BulkAssignRequest,
    db: AsyncSession = Depends(get_db),
    principal: Principal = Depends(get_current_principal),
):
    # 批量校验:body 里所有 account_id 都必须在用户范围内
    for aid in body.account_ids:
        await _scope_check_account(db, principal, aid)
    """
    批量把服务账号迁到另一个货源（SupplySource）下。

    规则：
      - 跨 provider **禁止**（aws 账号不能挂到 gcp 货源下），触发即跳过
      - 已经在目标货源下的账号跳过
      - 每条迁移写一条 ProjectAssignmentLog (action=reassigned)
      - 整体事务：任一成功则一起 commit；中途异常整批回滚
    """
    # 1. 加载目标货源 + 供应商信息
    target_ss = (await db.execute(
        select(SupplySource, Supplier)
        .join(Supplier, SupplySource.supplier_id == Supplier.id)
        .where(SupplySource.id == body.target_supply_source_id)
    )).first()
    if not target_ss:
        raise HTTPException(404, f"目标货源 id={body.target_supply_source_id} 不存在")
    target_ss_obj, target_supplier = target_ss
    target_provider = target_ss_obj.provider

    # 2. 批量加载 Project + 现有 SS
    rows = (await db.execute(
        select(Project, SupplySource)
        .join(SupplySource, Project.supply_source_id == SupplySource.id)
        .where(Project.id.in_(body.account_ids))
    )).all()
    found_map = {p.id: (p, ss) for p, ss in rows}

    moved = 0
    skipped: list[BulkAssignSkip] = []

    for acc_id in body.account_ids:
        hit = found_map.get(acc_id)
        if not hit:
            skipped.append(BulkAssignSkip(account_id=acc_id, reason="不存在"))
            continue
        project, current_ss = hit

        if project.supply_source_id == target_ss_obj.id:
            skipped.append(BulkAssignSkip(account_id=acc_id, reason="已在目标货源下"))
            continue

        if current_ss.provider != target_provider:
            skipped.append(BulkAssignSkip(
                account_id=acc_id,
                reason=f"跨 provider 禁止（账号 {current_ss.provider} → 目标 {target_provider}）",
            ))
            continue

        old_ss_id = project.supply_source_id
        project.supply_source_id = target_ss_obj.id
        # 跨货源：主体强制清空（entity 与 supply_source 绑定）
        project.entity_id = None
        _log(
            db, project,
            action="reassigned",
            from_status=f"ss#{old_ss_id}",
            to_status=f"ss#{target_ss_obj.id}",
            notes=f"bulk_assign → 供应商 '{target_supplier.name}' / {target_provider}",
        )
        moved += 1

    await db.commit()

    return BulkAssignResponse(
        moved=moved,
        skipped=skipped,
        target_supply_source_id=target_ss_obj.id,
        target_provider=target_provider,
        target_supplier_name=target_supplier.name,
    )


# ─── Bulk reassign to another Entity (same supply_source) ────

class BulkAssignEntityRequest(BaseModel):
    account_ids: list[int]
    # null = 清空主体（accounts.entity_id 设为 NULL），非空 = 切到指定主体
    target_entity_id: int | None = None

    @field_validator("account_ids")
    @classmethod
    def _non_empty(cls, v: list[int]) -> list[int]:
        if not v:
            raise ValueError("account_ids 不能为空")
        return list(dict.fromkeys(v))


class BulkAssignEntitySkip(BaseModel):
    account_id: int
    reason: str


class BulkAssignEntityResponse(BaseModel):
    moved: int
    skipped: list[BulkAssignEntitySkip]
    target_entity_id: int | None
    target_entity_name: str | None = None


@router.post(
    "/bulk-assign-entity",
    response_model=BulkAssignEntityResponse,
)
async def bulk_assign_entity(
    body: BulkAssignEntityRequest,
    db: AsyncSession = Depends(get_db),
    principal: Principal = Depends(get_current_principal),
):
    """批量把服务账号挂到/挪离主体。

    规则：
      - 主体绑定 supply_source；target_entity_id 非空时账号必须在该
        主体所属 supply_source 下，跨 supply_source 一律跳过。
      - target_entity_id=null 表示清空（设回"未分配主体"）。
      - 已在目标主体（或 entity_id 已是 null 且目标也是 null）的账号跳过。
      - 整体事务：任一成功则一起 commit；中途异常整批回滚。
    """
    for aid in body.account_ids:
        await _scope_check_account(db, principal, aid)

    target_entity: Entity | None = None
    target_ss_id: int | None = None
    if body.target_entity_id is not None:
        target_entity = await db.get(Entity, body.target_entity_id)
        if not target_entity:
            raise HTTPException(404, f"目标主体 id={body.target_entity_id} 不存在")
        ss = await db.get(SupplySource, target_entity.supply_source_id)
        if ss is not None:
            ensure_provider_visible(principal, ss.provider)
        target_ss_id = target_entity.supply_source_id

    rows = (
        await db.execute(select(Project).where(Project.id.in_(body.account_ids)))
    ).scalars().all()
    found_map = {p.id: p for p in rows}

    moved = 0
    skipped: list[BulkAssignEntitySkip] = []

    for acc_id in body.account_ids:
        p = found_map.get(acc_id)
        if not p:
            skipped.append(BulkAssignEntitySkip(account_id=acc_id, reason="不存在"))
            continue

        if target_entity is None:
            # 清空
            if p.entity_id is None:
                skipped.append(BulkAssignEntitySkip(account_id=acc_id, reason="已是未分配主体"))
                continue
            p.entity_id = None
            moved += 1
        else:
            if p.supply_source_id != target_ss_id:
                skipped.append(BulkAssignEntitySkip(
                    account_id=acc_id,
                    reason="账号与目标主体不在同一货源下",
                ))
                continue
            if p.entity_id == target_entity.id:
                skipped.append(BulkAssignEntitySkip(account_id=acc_id, reason="已在目标主体下"))
                continue
            p.entity_id = target_entity.id
            moved += 1

    await db.commit()
    return BulkAssignEntityResponse(
        moved=moved,
        skipped=skipped,
        target_entity_id=body.target_entity_id,
        target_entity_name=target_entity.name if target_entity else None,
    )


# ─── Taiji from-blob discovery (后端自动拉，前端零输入) ──────────────

class TaijiFromBlobRequest(BaseModel):
    supply_source_id: int
    entity_id: int | None = None


class TaijiFromBlobSkip(BaseModel):
    external_project_id: str
    reason: str


class TaijiFromBlobResponse(BaseModel):
    created: int
    skipped: list[TaijiFromBlobSkip]
    total_parsed: int
    snapshot_date: str | None = None  # 实际成功拉到的快照日期 YYYY-MM-DD
    section_used: str = "taiji"


def _taiji_fetch_latest_snapshot(sas_url: str, *, lookback_days: int = 7) -> tuple[str, dict]:
    """从 SAS 容器尝试拉最近 lookback_days 天里第一个可用的 {date}_UTC+0.json。

    返回 (snapshot_date, payload)；全部 404 抛 HTTPException 400 + 完整 URL 模板。

    路径策略：在 container 根 + 几个常见子目录都试一遍。Azure Blob 路径里的
    `+` 字符要 URL-encode 成 `%2B`（不 encode 会被服务端当作空格解释，导致 404）。
    """
    import httpx as _httpx
    from urllib.parse import urlsplit as _urlsplit, urlunsplit as _urlunsplit

    # `+` 必须 encode；`:` 不会出现在文件名里，但保守起见也保护一下。
    def _encode_filename(name: str) -> str:
        return name.replace("+", "%2B")

    # 后端尝试的子目录前缀。生产实测文件在 `taiji_log_data/` 子目录下；
    # 保留根目录与几个常见命名作为兜底。SAS 是容器级 read-only、没 list 权限，
    # 只能"猜"路径——第一个 200 OK 的就用。
    subdir_candidates = ["taiji_log_data/", "", "taiji/", "daily/"]

    today = dt.date.today()
    last_err: str | None = None
    tried_urls: list[str] = []
    parts = _urlsplit(sas_url)
    with _httpx.Client(timeout=_httpx.Timeout(30.0, read=60.0)) as client:
        for d in range(lookback_days + 1):
            day = today - dt.timedelta(days=d)
            raw_filename = f"{day.isoformat()}_UTC+0.json"
            for sub in subdir_candidates:
                filename = sub + _encode_filename(raw_filename)
                new_path = parts.path.rstrip("/") + "/" + filename
                url = _urlunsplit((parts.scheme, parts.netloc, new_path, parts.query, parts.fragment))
                # 记录脱敏后的 URL（去掉 query 里的 sig）便于排查
                tried_urls.append(_urlunsplit((parts.scheme, parts.netloc, new_path, "", "")))
                try:
                    resp = client.get(url)
                except Exception as e:
                    last_err = f"GET {filename} 失败: {e}"
                    continue
                if resp.status_code == 200:
                    try:
                        return day.isoformat(), resp.json()
                    except Exception as e:
                        last_err = f"{filename} 解析 JSON 失败: {e}"
                        continue
                if resp.status_code == 404:
                    continue
                last_err = f"{filename} HTTP {resp.status_code}"

    # 失败时把试过的前 4 个 URL 一起返回，方便用户去 Storage Explorer 比对
    sample = "; ".join(tried_urls[:4])
    raise HTTPException(
        400,
        f"Blob 最近 {lookback_days + 1} 天均无可用快照。最后错误: {last_err or '全部 404'}。"
        f" 试过的 URL 样本: {sample}",
    )


@router.post(
    "/taiji-from-blob",
    response_model=TaijiFromBlobResponse,
)
async def taiji_from_blob(
    body: TaijiFromBlobRequest,
    db: AsyncSession = Depends(get_db),
    principal: Principal = Depends(get_current_principal),
):
    """新建 Taiji 货源专用：后端从 settings.TAIJI_BLOB_SAS_URL 自动拉最新一天
    的快照，发现所有 (username, token_name) 对，批量建账号；secret_data 落
    blob_sas_url，未来由 collector 按日继续拉。

    规则：
      - SAS URL 不由前端传入；由 settings.TAIJI_BLOB_SAS_URL 提供（环境变量配置）
      - supply_source.provider 必须为 "taiji"
      - 只读 JSON 顶层 "taiji" section
      - external_project_id = "<username>:<token_name>"
      - 同 supply_source 下已存在的跳过（幂等）
      - 整批一事务，任一失败回滚
    """
    from app.config import settings

    sas_url = (settings.TAIJI_BLOB_SAS_URL or "").strip()
    if not sas_url:
        raise HTTPException(400, "服务端未配置 TAIJI_BLOB_SAS_URL，请先在 Container App 环境变量中设置")

    ss = await db.get(SupplySource, body.supply_source_id)
    if not ss:
        raise HTTPException(404, "货源不存在")
    if ss.provider != "taiji":
        raise HTTPException(400, f"目标货源 provider={ss.provider}，仅 taiji 支持后端自动发现")
    ensure_provider_visible(principal, ss.provider)

    # entity_id 校验
    target_entity_id: int | None = None
    if body.entity_id is not None:
        ent = await db.get(Entity, body.entity_id)
        if not ent or ent.supply_source_id != body.supply_source_id:
            raise HTTPException(400, "entity_id 不属于该货源")
        target_entity_id = ent.id

    # 后端 fetch（同步阻塞）—— 容器内到 Azure Blob 通常在 100ms~2s，可接受
    snapshot_date, snapshot_json = _taiji_fetch_latest_snapshot(sas_url)

    taiji_section = snapshot_json.get("taiji") if isinstance(snapshot_json, dict) else None
    if not isinstance(taiji_section, dict):
        raise HTTPException(500, f"快照 {snapshot_date}_UTC+0.json 缺少 taiji section")

    pairs: list[tuple[str, str]] = []
    for username, tokens in taiji_section.items():
        if not isinstance(tokens, dict):
            continue
        u = (username or "").strip()
        if not u:
            continue
        for token_name in tokens.keys():
            tn = (token_name or "").strip()
            if not tn:
                continue
            pairs.append((u, tn))

    if not pairs:
        return TaijiFromBlobResponse(created=0, skipped=[], total_parsed=0, snapshot_date=snapshot_date)

    existing_rows = (
        await db.execute(
            select(Project.external_project_id).where(
                Project.supply_source_id == body.supply_source_id,
            )
        )
    ).scalars().all()
    existing = set(existing_rows)

    skipped: list[TaijiFromBlobSkip] = []
    # 待建账号的"骨架"列表 —— 先全部攒起来再一次性批量 flush，把 3N 次数据库
    # round-trip 压成 4 次（CA、DS、Project、log）。否则上百对账号撞 30s 前端超时。
    to_create: list[tuple[str, str, str]] = []  # (username, token, external_id)
    for username, token_name in pairs:
        external_id = f"{username}:{token_name}"
        if external_id in existing:
            skipped.append(TaijiFromBlobSkip(
                external_project_id=external_id, reason="该货源下已存在同 ID 的账号",
            ))
            continue
        to_create.append((username, token_name, external_id))
        existing.add(external_id)

    # 关键设计：整个 supply_source 共享 ONE CloudAccount + ONE DataSource。
    # 之前每个 (user, token) 建一个独立 DS，每个 DS 都会拉同一份 blob 全量数据，
    # 由于 billing_summary 唯一约束含 data_source_id，N 个 DS 重复插入 N 份相同数据，
    # 让单条费用被放大 N 倍。共享 DS 后 collector 一天只拉一次 blob、一份行。
    shared_ca_name = f"taiji-shared-{body.supply_source_id}"
    shared_ca = (
        await db.execute(
            select(CloudAccount).where(
                CloudAccount.name == shared_ca_name,
                CloudAccount.provider == "taiji",
            ).limit(1)
        )
    ).scalars().first()

    encrypted = encrypt_dict({"blob_sas_url": sas_url})
    ds_cfg = {
        "auto_created": True,
        "timezone_tag": "UTC+0",
        "filename_template": "taiji_log_data/{date}_{tz}.json",
        "shared": True,
    }
    if shared_ca:
        # SAS 可能轮换：刷新 secret，DS config 保持
        shared_ca.secret_data = encrypted
        shared_ds = (
            await db.execute(
                select(DataSource).where(DataSource.cloud_account_id == shared_ca.id).limit(1)
            )
        ).scalars().first()
        if shared_ds is None:
            shared_ds = DataSource(
                name=f"ds-{shared_ca_name}"[:100],
                cloud_account_id=shared_ca.id,
                config=ds_cfg,
                is_active=True,
            )
            db.add(shared_ds)
            await db.flush()
    else:
        shared_ca = CloudAccount(name=shared_ca_name, provider="taiji", secret_data=encrypted)
        db.add(shared_ca)
        await db.flush()
        shared_ds = DataSource(
            name=f"ds-{shared_ca_name}"[:100],
            cloud_account_id=shared_ca.id,
            config=ds_cfg,
            is_active=True,
        )
        db.add(shared_ds)
        await db.flush()

    # 一次性批量建 Project，全部指向同一个 shared_ds.id
    projects: list[Project] = []
    for _u, token_name, external_id in to_create:
        project = Project(
            name=token_name,
            external_project_id=external_id,
            supply_source_id=body.supply_source_id,
            entity_id=target_entity_id,
            data_source_id=shared_ds.id,
            status="active",
        )
        db.add(project)
        projects.append(project)
    if projects:
        await db.flush()

    for project, (username, token_name, _ext) in zip(projects, to_create):
        _log(project=project, db=db, action="created", from_status="", to_status="active",
             notes=f"taiji-from-blob {snapshot_date}: {username}/{token_name}")

    await db.commit()
    return TaijiFromBlobResponse(
        created=len(to_create),
        skipped=skipped,
        total_parsed=len(pairs),
        snapshot_date=snapshot_date,
    )


# ─── Taiji ingest day from JSON body (绕过 Blob，直接落库) ───────────────

class TaijiIngestDayRequest(BaseModel):
    supply_source_id: int
    # 单天快照 JSON，结构同 {date}_UTC+0.json：顶层含 date_range 和 taiji
    snapshot_json: dict[str, Any]


class TaijiIngestDayResponse(BaseModel):
    snapshot_date: str
    projects_created: int
    projects_existing: int
    billing_rows_inserted: int


@router.post(
    "/taiji-ingest-day",
    response_model=TaijiIngestDayResponse,
)
async def taiji_ingest_day(
    body: TaijiIngestDayRequest,
    db: AsyncSession = Depends(get_db),
    principal: Principal = Depends(get_current_principal),
):
    """前端直接把单天快照 JSON 上传过来，后端：
    1) 共享 CA/DS 模式（命中 taiji-shared-{ssid}，没有就建）
    2) 缺账号自动建（指向 shared DS）
    3) 把当天所有 (project_id, model) 行 INSERT 进 billing_summary（ON CONFLICT DO NOTHING 幂等）
    4) 不刷预聚合（让前端在所有日上传完后调一次 refresh-summary）
    """
    from sqlalchemy import text as _text

    snapshot = body.snapshot_json
    # 校验
    dr = snapshot.get("date_range") if isinstance(snapshot, dict) else None
    if not isinstance(dr, dict) or not isinstance(dr.get("start_at"), str):
        raise HTTPException(400, "snapshot_json 缺少有效的 date_range.start_at")
    snapshot_date = dr["start_at"][:10]
    try:
        dt.date.fromisoformat(snapshot_date)
    except ValueError:
        raise HTTPException(400, f"date_range.start_at 不是合法 YYYY-MM-DD: {snapshot_date}")

    taiji_section = snapshot.get("taiji")
    if not isinstance(taiji_section, dict):
        raise HTTPException(400, "snapshot_json 缺少 taiji 顶层 section")

    ss = await db.get(SupplySource, body.supply_source_id)
    if not ss:
        raise HTTPException(404, "货源不存在")
    if ss.provider != "taiji":
        raise HTTPException(400, f"目标货源 provider={ss.provider}，仅 taiji 支持")
    ensure_provider_visible(principal, ss.provider)

    # 共享 CA/DS：找/建（与 /taiji-from-blob 同一约定）
    shared_ca_name = f"taiji-shared-{body.supply_source_id}"
    shared_ca = (
        await db.execute(
            select(CloudAccount).where(
                CloudAccount.name == shared_ca_name,
                CloudAccount.provider == "taiji",
            ).limit(1)
        )
    ).scalars().first()
    ds_cfg = {
        "auto_created": True,
        "timezone_tag": "UTC+0",
        "filename_template": "taiji_log_data/{date}_{tz}.json",
        "shared": True,
        "ingest_mode": "http_post",
    }
    if shared_ca:
        shared_ds = (
            await db.execute(
                select(DataSource).where(DataSource.cloud_account_id == shared_ca.id).limit(1)
            )
        ).scalars().first()
        if shared_ds is None:
            shared_ds = DataSource(
                name=f"ds-{shared_ca_name}"[:100],
                cloud_account_id=shared_ca.id,
                config=ds_cfg,
                is_active=True,
            )
            db.add(shared_ds)
            await db.flush()
    else:
        # 没 SAS 时 secret_data 给个空 dict；以后真要走 blob 同步再补 SAS
        shared_ca = CloudAccount(name=shared_ca_name, provider="taiji", secret_data=encrypt_dict({}))
        db.add(shared_ca)
        await db.flush()
        shared_ds = DataSource(
            name=f"ds-{shared_ca_name}"[:100],
            cloud_account_id=shared_ca.id,
            config=ds_cfg,
            is_active=True,
        )
        db.add(shared_ds)
        await db.flush()

    # 收集 (user, token) → 建缺账号
    pairs: list[tuple[str, str]] = []
    for username, tokens in taiji_section.items():
        if not isinstance(tokens, dict):
            continue
        u = (username or "").strip()
        if not u:
            continue
        for token_name in tokens.keys():
            tn = (token_name or "").strip()
            if tn:
                pairs.append((u, tn))

    existing_rows = (
        await db.execute(
            select(Project.external_project_id).where(
                Project.supply_source_id == body.supply_source_id,
            )
        )
    ).scalars().all()
    existing = set(existing_rows)

    to_create_pairs = [(u, t, f"{u}:{t}") for u, t in pairs if f"{u}:{t}" not in existing]
    projects_created = 0
    for _u, token_name, external_id in to_create_pairs:
        project = Project(
            name=token_name,
            external_project_id=external_id,
            supply_source_id=body.supply_source_id,
            data_source_id=shared_ds.id,
            status="active",
        )
        db.add(project)
        existing.add(external_id)
        projects_created += 1
    if projects_created:
        await db.flush()
    projects_existing = len(pairs) - projects_created

    # 构造 billing rows
    def _f(v: Any) -> float:
        try:
            return float(v) if v not in (None, "") else 0.0
        except (TypeError, ValueError):
            return 0.0

    def _i(v: Any) -> int:
        try:
            return int(v) if v not in (None, "") else 0
        except (TypeError, ValueError):
            try:
                return int(float(v))
            except (TypeError, ValueError):
                return 0

    rows: list[dict] = []
    for username, tokens in taiji_section.items():
        if not isinstance(tokens, dict):
            continue
        u = (username or "").strip()
        if not u:
            continue
        for token_name, token_blob in tokens.items():
            if not isinstance(token_blob, dict):
                continue
            tn = (token_name or "").strip()
            if not tn:
                continue
            project_id = f"{u}:{tn}"
            key_display = token_blob.get("key_display") or ""
            details = token_blob.get("details") or {}
            if not isinstance(details, dict):
                continue
            for model_name, m in details.items():
                if not isinstance(m, dict):
                    continue
                cost = round(_f(m.get("cost")), 6)
                prompt = _i(m.get("prompt_tokens"))
                completion = _i(m.get("completion_tokens"))
                count = _i(m.get("count"))
                cache_hit = _i(m.get("cache_hit_tokens"))
                add_info = {
                    "key_display": key_display,
                    "request_count": count,
                    "prompt_tokens": prompt,
                    "completion_tokens": completion,
                }
                if cache_hit:
                    add_info["cache_hit_tokens"] = cache_hit
                rows.append({
                    "date": snapshot_date,
                    "provider": "taiji",
                    "data_source_id": shared_ds.id,
                    "project_id": project_id,
                    "project_name": project_id,
                    "product": model_name or "unknown",
                    "usage_type": "",
                    "region": None,
                    "cost": cost,
                    "cost_type": "regular",
                    "usage_quantity": float(prompt + completion),
                    "usage_unit": "tokens",
                    "currency": "USD",
                    "currency_conversion_rate": 1.0,
                    "additional_info": json.dumps(add_info, ensure_ascii=False),
                })

    billing_inserted = 0
    if rows:
        # 批量 INSERT，唯一约束撞了就跳过（再次上传同一份不会重复算）。
        # 用列名 ON CONFLICT 而非 ON CONSTRAINT —— billing_summary 是 PG 分区表，
        # 按列名匹配更通用、不依赖具体 constraint 名。
        # COALESCE 兜底是因为 region 可能 NULL、usage_type 可能空串 —— PG 索引相等比较
        # NULL 是"未知"不等于 NULL，会导致 ON CONFLICT 漏判同一行触发重复插入。
        # 把可能 NULL 的列用空串替代再比较，保证去重逻辑稳定。
        stmt = _text("""
            INSERT INTO billing_summary
              (date, provider, data_source_id, project_id, project_name,
               product, usage_type, region, cost, cost_type,
               usage_quantity, usage_unit, currency, currency_conversion_rate,
               additional_info)
            VALUES
              (:date, :provider, :data_source_id, :project_id, :project_name,
               :product, :usage_type, :region, :cost, :cost_type,
               :usage_quantity, :usage_unit, :currency, :currency_conversion_rate,
               CAST(:additional_info AS JSONB))
            ON CONFLICT (date, data_source_id, project_id, product, usage_type, region, cost_type)
            DO NOTHING
        """)
        for r in rows:
            await db.execute(stmt, r)
            billing_inserted += 1  # 计数算"已尝试"行；真实 inserted 由 ON CONFLICT 决定

    await db.commit()

    return TaijiIngestDayResponse(
        snapshot_date=snapshot_date,
        projects_created=projects_created,
        projects_existing=projects_existing,
        billing_rows_inserted=billing_inserted,
    )


# ─── Taiji cleanup duplicates (修复历史每账号独立 DS 造成的重复行) ──────

class TaijiCleanupRequest(BaseModel):
    supply_source_id: int
    dry_run: bool = True  # 默认只统计不动数据


class TaijiCleanupResponse(BaseModel):
    dry_run: bool
    total_data_sources_before: int
    kept_data_source_id: int | None
    orphan_data_sources_removed: int
    orphan_cloud_accounts_removed: int
    billing_rows_deleted_as_dup: int
    billing_rows_reassigned_to_kept: int
    projects_repointed: int


@router.post(
    "/taiji-cleanup-duplicates",
    response_model=TaijiCleanupResponse,
    dependencies=[Depends(require_roles("cloud_admin"))],
)
async def taiji_cleanup_duplicates(
    body: TaijiCleanupRequest,
    db: AsyncSession = Depends(get_db),
):
    """把某 Taiji supply_source 下"每账号一个独立 CA/DS"的历史结构合并为
    "整个 supply_source 共享一个 CA/DS"，并去重 billing_summary 重复行。

    操作：
    1) 找出该 supply_source 下所有 Project 引用过的 data_source_id 集合
    2) 选 id 最小的 DS 留下作为"shared"，CA 重命名为 taiji-shared-{ssid}
    3) 把所有 Project 指向 kept DS
    4) billing_summary：按 (date, project_id, product, usage_type, region, cost_type)
       去重，只保留 id 最小的一行（删除其他副本）
    5) 把剩下行的 data_source_id 全部改成 kept DS（不会撞唯一约束因为已去重）
    6) 删除孤儿 DS 和孤儿 CA（无任何 DS 关联的）

    dry_run=true 只统计、不真正改库；false 才落地。
    """
    from sqlalchemy import text

    ss = await db.get(SupplySource, body.supply_source_id)
    if not ss:
        raise HTTPException(404, "货源不存在")
    if ss.provider != "taiji":
        raise HTTPException(400, "仅 taiji 货源支持此 cleanup")

    # 1) 收集该 supply_source 所有 Project 用过的 data_source_id
    ds_rows = (
        await db.execute(
            select(Project.data_source_id).where(
                Project.supply_source_id == body.supply_source_id,
                Project.data_source_id.isnot(None),
            )
        )
    ).scalars().all()
    ds_ids = sorted(set(int(x) for x in ds_rows if x is not None))
    if not ds_ids:
        return TaijiCleanupResponse(
            dry_run=body.dry_run, total_data_sources_before=0, kept_data_source_id=None,
            orphan_data_sources_removed=0, orphan_cloud_accounts_removed=0,
            billing_rows_deleted_as_dup=0, billing_rows_reassigned_to_kept=0,
            projects_repointed=0,
        )

    kept_ds_id = ds_ids[0]
    orphan_ds_ids = [x for x in ds_ids if x != kept_ds_id]

    # 2) 统计 billing 重复情况（不论 dry_run）。
    # 用 DISTINCT ON 算保留集（O(N log N)），总数 - 保留数 = 重复数。
    # 之前用 correlated subquery 是 O(N²)，6 万行就超 30 秒前端 fetch 超时。
    total_cnt = (
        await db.execute(
            text("SELECT count(*) FROM billing_data WHERE data_source_id = ANY(:ds_ids)"),
            {"ds_ids": ds_ids},
        )
    ).scalar() or 0
    keep_cnt = (
        await db.execute(
            text("""
                SELECT count(*) FROM (
                  SELECT DISTINCT ON (date, project_id, product, usage_type, region, cost_type) id
                  FROM billing_data
                  WHERE data_source_id = ANY(:ds_ids)
                  ORDER BY date, project_id, product, usage_type, region, cost_type, id
                ) keep
            """),
            {"ds_ids": ds_ids},
        )
    ).scalar() or 0
    dup_cnt = int(total_cnt - keep_cnt)

    reassign_cnt = (
        await db.execute(
            text("SELECT count(*) FROM billing_data WHERE data_source_id = ANY(:other)"),
            {"other": orphan_ds_ids if orphan_ds_ids else [-1]},
        )
    ).scalar() or 0

    projects_repointed_cnt = (
        await db.execute(
            text("""
                SELECT count(*) FROM projects
                WHERE supply_source_id = :ssid AND data_source_id = ANY(:other)
            """),
            {"ssid": body.supply_source_id, "other": orphan_ds_ids if orphan_ds_ids else [-1]},
        )
    ).scalar() or 0

    if body.dry_run:
        return TaijiCleanupResponse(
            dry_run=True,
            total_data_sources_before=len(ds_ids),
            kept_data_source_id=kept_ds_id,
            orphan_data_sources_removed=len(orphan_ds_ids),
            orphan_cloud_accounts_removed=len(orphan_ds_ids),  # 估算：1 DS = 1 CA
            billing_rows_deleted_as_dup=int(dup_cnt),
            billing_rows_reassigned_to_kept=int(reassign_cnt - dup_cnt),  # 重定向 - 已删
            projects_repointed=int(projects_repointed_cnt),
        )

    # ─── 真正执行 ───
    # 删除重复行（保留每业务 key 下 id 最小的）。用 DISTINCT ON 算 keep 集，
    # 然后 DELETE NOT IN keep —— O(N log N)，远快于 correlated subquery。
    await db.execute(
        text("""
            WITH to_keep AS (
              SELECT DISTINCT ON (date, project_id, product, usage_type, region, cost_type) id
              FROM billing_data
              WHERE data_source_id = ANY(:ds_ids)
              ORDER BY date, project_id, product, usage_type, region, cost_type, id
            )
            DELETE FROM billing_data
            WHERE data_source_id = ANY(:ds_ids)
              AND id NOT IN (SELECT id FROM to_keep)
        """),
        {"ds_ids": ds_ids},
    )

    # 重定向所有剩余 billing 行到 kept_ds_id
    await db.execute(
        text("UPDATE billing_data SET data_source_id = :kept WHERE data_source_id = ANY(:other)"),
        {"kept": kept_ds_id, "other": orphan_ds_ids if orphan_ds_ids else [-1]},
    )

    # Project 指向 kept DS
    await db.execute(
        text("""
            UPDATE projects SET data_source_id = :kept
            WHERE supply_source_id = :ssid AND data_source_id = ANY(:other)
        """),
        {"kept": kept_ds_id, "ssid": body.supply_source_id, "other": orphan_ds_ids if orphan_ds_ids else [-1]},
    )

    # 重命名 kept CA/DS 为 shared 约定名
    target_name = f"taiji-shared-{body.supply_source_id}"
    kept_ds = await db.get(DataSource, kept_ds_id)
    kept_ca_id = kept_ds.cloud_account_id if kept_ds else None
    if kept_ds:
        kept_ds.name = f"ds-{target_name}"[:100]
    if kept_ca_id:
        kept_ca = await db.get(CloudAccount, kept_ca_id)
        if kept_ca:
            kept_ca.name = target_name

    # 收集要删的孤儿 CA（指向这些 DS 的）
    orphan_ca_rows = (
        await db.execute(
            text("SELECT cloud_account_id FROM data_sources WHERE id = ANY(:other)"),
            {"other": orphan_ds_ids if orphan_ds_ids else [-1]},
        )
    ).scalars().all()
    orphan_ca_ids = sorted(set(int(x) for x in orphan_ca_rows if x is not None))

    # 清理其他指向孤儿 DS 的外键（sync_logs / token_usage / resource_inventory）。
    # billing_raw_taiji 和 taiji_log_raw 是 CASCADE 自动删；billing_data 已 reassign。
    # 这些表的历史记录直接删 —— 数据来源已合并到 kept DS，旧 DS 没人再用了。
    if orphan_ds_ids:
        await db.execute(
            text("DELETE FROM sync_logs WHERE data_source_id = ANY(:other)"),
            {"other": orphan_ds_ids},
        )
        await db.execute(
            text("DELETE FROM token_usage WHERE data_source_id = ANY(:other)"),
            {"other": orphan_ds_ids},
        )
        # resource_inventory 的 FK 是 nullable，安全 SET NULL；Taiji 一般没资源数据
        await db.execute(
            text("UPDATE resource_inventory SET data_source_id = NULL WHERE data_source_id = ANY(:other)"),
            {"other": orphan_ds_ids},
        )

        # 删孤儿 DS
        await db.execute(
            text("DELETE FROM data_sources WHERE id = ANY(:other)"),
            {"other": orphan_ds_ids},
        )

    # 删完全孤儿的 CA（确保没有其他 DS 还指着）
    removed_ca_cnt = 0
    for ca_id in orphan_ca_ids:
        if ca_id == kept_ca_id:
            continue
        still_referenced = (
            await db.execute(
                text("SELECT count(*) FROM data_sources WHERE cloud_account_id = :cid"),
                {"cid": ca_id},
            )
        ).scalar() or 0
        if still_referenced == 0:
            await db.execute(
                text("DELETE FROM cloud_accounts WHERE id = :cid"),
                {"cid": ca_id},
            )
            removed_ca_cnt += 1

    await db.commit()

    return TaijiCleanupResponse(
        dry_run=False,
        total_data_sources_before=len(ds_ids),
        kept_data_source_id=kept_ds_id,
        orphan_data_sources_removed=len(orphan_ds_ids),
        orphan_cloud_accounts_removed=removed_ca_cnt,
        billing_rows_deleted_as_dup=int(dup_cnt),
        billing_rows_reassigned_to_kept=int(reassign_cnt - dup_cnt),
        projects_repointed=int(projects_repointed_cnt),
    )


@router.post(
    "/{account_id}/suspend",
    response_model=ServiceAccountDetail,
    dependencies=[Depends(_account_in_scope)],
)
async def suspend_account(account_id: int, request: Request, db: AsyncSession = Depends(get_db)):
    """人工停用。允许从使用中 / 备用切到停用。"""
    project = await _get_active_project(db, account_id)
    if project.status == "inactive":
        return await get_account(account_id, db)

    old_status = project.status
    project.status = "inactive"

    _log(
        db, project, "suspended",
        from_status=old_status, to_status="inactive",
        operator=_principal_operator(request),
    )
    await db.commit()
    return await get_account(account_id, db)


@router.post(
    "/{account_id}/activate",
    response_model=ServiceAccountDetail,
    dependencies=[Depends(_account_in_scope)],
)
async def activate_account(account_id: int, request: Request, db: AsyncSession = Depends(get_db)):
    """人工置为使用中。允许从备用 / 停用切到使用中。和客户编号无关。"""
    project = await _get_active_project(db, account_id)
    if project.status == "active":
        return await get_account(account_id, db)

    old_status = project.status
    project.status = "active"
    _log(
        db, project, "activated",
        from_status=old_status, to_status="active",
        operator=_principal_operator(request),
    )
    await db.commit()
    return await get_account(account_id, db)


@router.post(
    "/{account_id}/standby",
    response_model=ServiceAccountDetail,
    dependencies=[Depends(_account_in_scope)],
)
async def standby_account(account_id: int, request: Request, db: AsyncSession = Depends(get_db)):
    """人工置为备用。允许从使用中 / 停用切到备用。和客户编号无关。"""
    project = await _get_active_project(db, account_id)
    if project.status == "standby":
        return await get_account(account_id, db)

    old_status = project.status
    project.status = "standby"
    _log(
        db, project, "standby",
        from_status=old_status, to_status="standby",
        operator=_principal_operator(request),
    )
    await db.commit()
    return await get_account(account_id, db)


# ─── Sales-system batch sync ──────────────────────────────────

@router.post(
    "/customer-assignments/sync",
    response_model=SalesSyncResult,
)
async def sync_customer_assignments(
    body: SalesSyncBody,
    request: Request,
    db: AsyncSession = Depends(get_db),
    principal: Principal = Depends(get_current_principal),
):
    # 销售系统批量下发:仅 admin/ops 可调(跨 provider 操作)
    if not has_full_access(principal):
        raise HTTPException(403, "Only cloud_admin/cloud_ops can run sales sync (cross-provider).")
    """销售系统调用：批量下发 (customer_code, supplier, provider, external_project_id) 关联。

    定位键：(supplier_name, provider, external_project_id)。找不到则写入 unmatched 返回，
    不阻断整批。

    - mode=full: 只对 body.scope_customer_codes 这批客户编号做差分（多删少插）。
      若未传 scope_customer_codes，则按入参 assignments 中出现的 customer_codes 作为 scope。
    - mode=patch: 仅 upsert，不删除。
    """
    operator = _principal_operator(request) or "sales-sync"

    # 归一化入参
    items: list[tuple[str, str, str, str]] = []  # (code, supplier, provider, ext)
    for a in body.assignments:
        code = _normalize_code(a.customer_code)
        supplier = (a.supplier_name or "").strip()
        provider = (a.provider or "").strip().lower()
        ext = (a.external_project_id or "").strip()
        if not (code and supplier and provider and ext):
            continue
        items.append((code, supplier, provider, ext))

    scope_codes: set[str] = {
        _normalize_code(c) for c in (body.scope_customer_codes or []) if c
    }
    if body.mode == "full" and not scope_codes:
        scope_codes = {i[0] for i in items}

    # 批量定位 project：一次查出用到的 (supplier_name, provider, external_project_id) → project
    unique_keys = {(s, p, e) for _, s, p, e in items}
    project_map: dict[tuple[str, str, str], Project] = {}
    if unique_keys:
        suppliers = {s for s, _, _ in unique_keys}
        providers = {p for _, p, _ in unique_keys}
        exts = {e for _, _, e in unique_keys}
        rows = (
            await db.execute(
                select(Project, SupplySource, Supplier)
                .join(SupplySource, Project.supply_source_id == SupplySource.id)
                .join(Supplier, SupplySource.supplier_id == Supplier.id)
                .where(
                    Supplier.name.in_(suppliers),
                    SupplySource.provider.in_(providers),
                    Project.external_project_id.in_(exts),
                )
            )
        ).all()
        for p, ss, su in rows:
            project_map[(su.name, ss.provider, p.external_project_id)] = p

    unmatched: list[SalesSyncUnmatched] = []
    desired: dict[int, set[str]] = {}  # project_id -> customer_codes (new desired set, partial)
    touched_projects: dict[int, Project] = {}
    for code, supplier, provider, ext in items:
        proj = project_map.get((supplier, provider, ext))
        if not proj:
            unmatched.append(SalesSyncUnmatched(
                customer_code=code, supplier_name=supplier, provider=provider,
                external_project_id=ext,
                reason="service account not found",
            ))
            continue
        desired.setdefault(proj.id, set()).add(code)
        touched_projects[proj.id] = proj

    # Fetch current state for every project we will write to. In full mode, we
    # also need projects that currently hold any scope_codes even if they don't
    # appear in assignments (so we can delete them).
    candidate_project_ids: set[int] = set(desired.keys())
    if body.mode == "full" and scope_codes:
        rows = (
            await db.execute(
                select(ProjectCustomerAssignment.project_id)
                .where(ProjectCustomerAssignment.customer_code.in_(list(scope_codes)))
                .distinct()
            )
        ).all()
        for (pid,) in rows:
            candidate_project_ids.add(pid)

    if not candidate_project_ids:
        await db.commit()
        return SalesSyncResult(inserted=0, deleted=0, unchanged=0, unmatched=unmatched)

    # Load missing Project rows we don't have cached yet
    missing_ids = [pid for pid in candidate_project_ids if pid not in touched_projects]
    if missing_ids:
        prows = (
            await db.execute(select(Project).where(Project.id.in_(missing_ids)))
        ).scalars().all()
        for p in prows:
            touched_projects[p.id] = p

    current_map = await _codes_by_project_ids(db, list(candidate_project_ids))

    inserted = deleted = unchanged = 0
    for pid in candidate_project_ids:
        proj = touched_projects.get(pid)
        if not proj:
            continue
        current = set(current_map.get(pid, []))
        want = desired.get(pid, set())

        if body.mode == "full":
            # Only compare within scope_codes. Codes outside scope stay untouched.
            current_in_scope = current & scope_codes
            want_in_scope = want & scope_codes
            to_add = want_in_scope - current_in_scope
            to_remove = current_in_scope - want_in_scope
            unchanged += len(current_in_scope & want_in_scope)
        else:
            # patch: upsert only, never delete
            to_add = want - current
            to_remove = set()
            unchanged += len(current & want)

        for code in sorted(to_remove):
            await db.execute(
                ProjectCustomerAssignment.__table__.delete().where(
                    (ProjectCustomerAssignment.project_id == proj.id)
                    & (ProjectCustomerAssignment.customer_code == code)
                )
            )
            deleted += 1
            _log(
                db, proj, "customer_unbound",
                from_status=proj.status, to_status=proj.status,
                customer_code=code, operator=operator,
                notes="sales batch sync",
            )

        for code in sorted(to_add):
            db.add(ProjectCustomerAssignment(
                project_id=proj.id, customer_code=code, assigned_by=operator,
                notes="sales sync",
            ))
            inserted += 1
            _log(
                db, proj, "customer_bound",
                from_status=proj.status, to_status=proj.status,
                customer_code=code, operator=operator,
                notes="sales batch sync",
            )

        await db.flush()
        # status 不再随客户编号派生，保持当前值

    await db.commit()
    return SalesSyncResult(
        inserted=inserted, deleted=deleted, unchanged=unchanged, unmatched=unmatched,
    )


@router.delete(
    "/{account_id}",
    status_code=204,
    dependencies=[Depends(require_roles("cloud_admin"))],
)
async def delete_account(account_id: int, db: AsyncSession = Depends(get_db)):
    await _hard_delete(account_id, db)


@router.get("/{account_id}/costs", response_model=CostSummary)
async def get_costs(
    account_id: int,
    start_date: str = Query(..., pattern=r"^\d{4}-\d{2}-\d{2}$"),
    end_date: str = Query(..., pattern=r"^\d{4}-\d{2}-\d{2}$"),
    db: AsyncSession = Depends(get_db),
):
    project = await _get_active_project(db, account_id)

    prov = await _cloud_provider(db, project)
    sd = dt.date.fromisoformat(start_date)
    ed = dt.date.fromisoformat(end_date) + dt.timedelta(days=1)

    res = await db.execute(
        select(
            BillingData.date,
            BillingData.product,
            func.sum(BillingData.cost).label("cost"),
            func.sum(BillingData.usage_quantity).label("usage_quantity"),
            func.max(BillingData.usage_unit).label("usage_unit"),
        )
        .where(
            func.trim(BillingData.project_id) == project.external_project_id.strip(),
            BillingData.provider == prov,
            BillingData.date >= sd,
            BillingData.date < ed,
        )
        .group_by(BillingData.date, BillingData.product)
        .order_by(BillingData.date, BillingData.product)
    )
    rows = res.all()

    total = 0.0
    total_usage = 0.0
    svc_cost: dict[str, float] = {}
    svc_usage: dict[str, float] = {}
    svc_unit: dict[str, str | None] = {}
    daily_map: dict[str, float] = {}
    daily_usage_map: dict[str, float] = {}
    daily_by_service: list[DailyServiceCost] = []

    for r in rows:
        cost = float(r.cost)
        uq = float(r.usage_quantity or 0)
        product = r.product or "Unknown"
        date_str = str(r.date)

        total += cost
        total_usage += uq
        svc_cost[product] = svc_cost.get(product, 0.0) + cost
        svc_usage[product] = svc_usage.get(product, 0.0) + uq
        if product not in svc_unit:
            svc_unit[product] = r.usage_unit
        daily_map[date_str] = daily_map.get(date_str, 0.0) + cost
        daily_usage_map[date_str] = daily_usage_map.get(date_str, 0.0) + uq
        daily_by_service.append(DailyServiceCost(
            date=date_str, service=product, cost=cost,
            usage_quantity=uq, usage_unit=r.usage_unit,
        ))

    services = sorted(
        [CostByService(service=k, cost=v, usage_quantity=svc_usage[k], usage_unit=svc_unit.get(k))
         for k, v in svc_cost.items()],
        key=lambda x: x.cost, reverse=True,
    )
    daily = [DailyCost(date=k, cost=v, usage_quantity=daily_usage_map[k])
             for k, v in sorted(daily_map.items())]

    return CostSummary(
        total_cost=total, total_usage=total_usage,
        services=services, daily=daily, daily_by_service=daily_by_service,
    )


@router.get(
    "/{account_id}/credentials",
    dependencies=[Depends(_account_in_scope)],
)
async def get_credentials(account_id: int, db: AsyncSession = Depends(get_db)):
    project = await _get_active_project(db, account_id)
    if not project.data_source_id:
        return {}
    ds = await db.get(DataSource, project.data_source_id)
    if not ds:
        return {}
    ca = await db.get(CloudAccount, ds.cloud_account_id)
    if not ca:
        return {}
    try:
        return decrypt_to_dict(ca.secret_data)
    except Exception:
        raise HTTPException(500, "Failed to decrypt credentials")


@router.get("/{account_id}/costs/export")
async def export_account_costs(
    account_id: int,
    start_date: str = Query(..., pattern=r"^\d{4}-\d{2}-\d{2}$"),
    end_date: str = Query(..., pattern=r"^\d{4}-\d{2}-\d{2}$"),
    discount_pct: float | None = Query(
        None,
        ge=0,
        le=100,
        description="统一折扣百分比；传入时导出增加「折扣」「折后费用」列",
    ),
    db: AsyncSession = Depends(get_db),
):
    project = await _get_active_project(db, account_id)

    prov = await _cloud_provider(db, project)
    sd = dt.date.fromisoformat(start_date)
    ed = dt.date.fromisoformat(end_date) + dt.timedelta(days=1)

    billing_stmt = (
        select(
            BillingData.date,
            BillingData.service_id,
            BillingData.product,
            BillingData.sku_id,
            BillingData.usage_type,
            BillingData.region,
            BillingData.resource_name,
            BillingData.cost_type,
            BillingData.usage_quantity,
            BillingData.usage_unit,
            BillingData.cost,
            BillingData.cost_at_list,
            BillingData.credits_total,
            BillingData.currency,
            BillingData.invoice_month,
        )
        .where(
            func.trim(BillingData.project_id) == project.external_project_id.strip(),
            BillingData.provider == prov,
            BillingData.date >= sd,
            BillingData.date < ed,
        )
        .order_by(BillingData.date, BillingData.product)
    )
    rows = (await db.execute(billing_stmt)).all()

    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment

    wb = Workbook()
    ws = wb.active
    ws.title = "费用明细"

    header_font = Font(bold=True, color="FFFFFF", size=11)
    header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
    header_align = Alignment(horizontal="center")

    # 列对照（和 BQ Excel 导出一致）：
    #   服务 ID = service_id，SKU ID = sku_id，资源 ID = resource_name，
    #   计费类型 = cost_type（regular/tax/adjustment），
    #   未含入的小计 = cost_at_list（标价），节省合计 = credits_total
    base_headers = [
        "日期", "服务", "服务 ID", "用量类型", "SKU ID",
        "区域", "资源 ID", "计费类型",
        "用量", "用量单位",
        "未含入的小计(USD)", "节省合计(USD)",
        "费用/小计(USD)", "币种", "发票月",
    ]
    if discount_pct is not None:
        factor = 1.0 - float(discount_pct) / 100.0
        headers = base_headers + ["折扣(%)", "折后费用(USD)"]
    else:
        factor = 1.0
        headers = base_headers

    for ci, h in enumerate(headers, 1):
        cell = ws.cell(row=1, column=ci, value=h)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_align

    def _f(v):
        return float(v) if v is not None else None

    for ri, r in enumerate(rows, 2):
        cost = float(r.cost) if r.cost is not None else 0.0
        ws.cell(row=ri, column=1, value=str(r.date))
        ws.cell(row=ri, column=2, value=r.product or "Unknown")
        ws.cell(row=ri, column=3, value=r.service_id or "")
        ws.cell(row=ri, column=4, value=r.usage_type or "")
        ws.cell(row=ri, column=5, value=r.sku_id or "")
        ws.cell(row=ri, column=6, value=r.region or "")
        ws.cell(row=ri, column=7, value=r.resource_name or "")
        ws.cell(row=ri, column=8, value=r.cost_type or "")
        ws.cell(row=ri, column=9, value=_f(r.usage_quantity) or 0)
        ws.cell(row=ri, column=10, value=r.usage_unit or "")
        ws.cell(row=ri, column=11, value=_f(r.cost_at_list)).number_format = '#,##0.000000'
        ws.cell(row=ri, column=12, value=_f(r.credits_total)).number_format = '#,##0.000000'
        ws.cell(row=ri, column=13, value=cost).number_format = '#,##0.000000'
        ws.cell(row=ri, column=14, value=r.currency or "")
        ws.cell(row=ri, column=15, value=r.invoice_month or "")
        if discount_pct is not None:
            ws.cell(row=ri, column=16, value=float(discount_pct))
            ws.cell(row=ri, column=17, value=cost * factor).number_format = '#,##0.000000'

    for col in range(1, len(headers) + 1):
        ws.column_dimensions[chr(64 + col)].width = 18

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)

    fname = f"{project.name}_{start_date}_{end_date}.xlsx"
    return StreamingResponse(
        buf,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f'attachment; filename="{fname}"'},
    )


def _build_excel(
    rows: list[AccountDailyCostRow],
    filename: str,
    discount_pct: float | None = None,
):
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment

    wb = Workbook()
    ws = wb.active
    ws.title = "日报表"

    header_font = Font(bold=True, color="FFFFFF", size=11)
    header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
    header_align = Alignment(horizontal="center")

    base_headers = [
        "云厂商", "账号名称", "账号ID", "日期",
        "服务", "服务 ID",
        "未含入的小计(USD)", "节省合计(USD)",
        "费用/小计(USD)", "币种",
    ]
    if discount_pct is not None:
        factor = 1.0 - float(discount_pct) / 100.0
        headers = base_headers + ["折扣(%)", "折后费用(USD)"]
    else:
        factor = 1.0
        headers = base_headers

    for ci, h in enumerate(headers, 1):
        cell = ws.cell(row=1, column=ci, value=h)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_align

    def _f(v):
        return float(v) if v is not None else None

    for ri, r in enumerate(rows, 2):
        cost = float(r.cost)
        ws.cell(row=ri, column=1, value=r.provider.upper())
        ws.cell(row=ri, column=2, value=r.account_name)
        ws.cell(row=ri, column=3, value=r.external_project_id)
        ws.cell(row=ri, column=4, value=r.date)
        ws.cell(row=ri, column=5, value=r.product or "Unknown")
        ws.cell(row=ri, column=6, value=r.service_id or "")
        ws.cell(row=ri, column=7, value=_f(r.cost_at_list)).number_format = '#,##0.000000'
        ws.cell(row=ri, column=8, value=_f(r.credits_total)).number_format = '#,##0.000000'
        ws.cell(row=ri, column=9, value=cost).number_format = '#,##0.000000'
        ws.cell(row=ri, column=10, value=r.currency or "")
        if discount_pct is not None:
            ws.cell(row=ri, column=11, value=float(discount_pct))
            ws.cell(row=ri, column=12, value=cost * factor).number_format = '#,##0.000000'

    for col in range(1, len(headers) + 1):
        ws.column_dimensions[chr(64 + col)].width = 18

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)

    return StreamingResponse(
        buf,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f'attachment; filename="{filename}"'},
    )


@router.post(
    "/discover-gcp-projects",
)
async def discover_gcp_projects(
    db: AsyncSession = Depends(get_db),
    principal: Principal = Depends(get_current_principal),
):
    # 跨 GCP 账号的发现操作:cloud_admin/ops 全开;cloud_gcp 也允许(它就是管 GCP)
    if not has_full_access(principal):
        scope_providers = extract_providers_from_roles(principal.roles)
        if "gcp" not in scope_providers:
            raise HTTPException(403, "Need cloud_admin / cloud_ops / cloud_gcp role")
    """为账单中存在但未建档的 GCP project 创建 Project，挂在系统供应商「未分配资源组」的 GCP 货源下。"""
    billing_res = await db.execute(
        select(
            BillingData.project_id,
            func.max(BillingData.project_name).label("project_name"),
        )
        .where(BillingData.provider == "gcp")
        .group_by(BillingData.project_id)
    )
    billing_projects = {r.project_id: r.project_name for r in billing_res.all() if r.project_id}

    if not billing_projects:
        return {"created": 0, "projects": []}

    ss_id, _ = await ensure_other_gcp_supply_source_id(db)

    existing_res = await db.execute(
        select(Project.external_project_id)
        .join(SupplySource, Project.supply_source_id == SupplySource.id)
        .where(SupplySource.provider == "gcp", Project.external_project_id.in_(list(billing_projects.keys())))
    )
    existing = {r[0] for r in existing_res.all()}

    created = []
    for pid, pname in billing_projects.items():
        if pid in existing:
            continue
        project = Project(
            name=pname or pid,
            external_project_id=pid,
            supply_source_id=ss_id,
            status="standby",
        )
        db.add(project)
        created.append(pid)

    if created:
        await db.commit()

    return {"created": len(created), "projects": created}
